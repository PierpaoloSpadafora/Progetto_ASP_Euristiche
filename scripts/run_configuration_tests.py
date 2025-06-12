
import clingo
import time
import statistics
import json
import multiprocessing
from pathlib import Path
import pandas as pd
import re

def load_config():
    with open("settings.json", "r", encoding="utf-8") as f:
        return json.load(f)

def cost_to_comparable(cost):
    if cost is None:
        return [float('inf')]
    if isinstance(cost, list):
        return cost
    return [cost]

def is_cost_better(cost1, cost2):

    if cost1 is None:
        return False
    if cost2 is None:
        return True
    
    comp1 = cost_to_comparable(cost1)
    comp2 = cost_to_comparable(cost2)
    
    for c1, c2 in zip(comp1, comp2):
        if c1 < c2:
            return True
        elif c1 > c2:
            return False
    
    return len(comp1) < len(comp2)

def run_clingo_worker(input_data, encoding, heuristics, control_args, timeout_seconds, shared_dict, lock):
    best_cost = None
    model_count = 0
    start_time = time.time()
    
    def on_model(model):
        nonlocal best_cost, model_count
        model_count += 1
        current_cost = model.cost
        
        if best_cost is None or is_cost_better(current_cost, best_cost):
            best_cost = current_cost
        
        with lock:
            shared_dict['best_cost'] = best_cost
            shared_dict['model_count'] = model_count
            shared_dict['elapsed_time'] = time.time() - start_time
        
        return True
    
    try:
        ctl = clingo.Control(control_args)
        ctl.add("base", [], input_data)
        ctl.add("base", [], encoding)
        
        if heuristics:
            ctl.add("base", [], heuristics)
        
        ctl.ground([("base", [])])
        
        result = ctl.solve(on_model=on_model)
        
        try:
            stats = ctl.statistics['solving']['solvers']
            solver_stats = {
                'choices': stats.get('choices', 0),
                'conflicts': stats.get('conflicts', 0)
            }
        except:
            solver_stats = {'choices': 0, 'conflicts': 0}
        
        elapsed_time = time.time() - start_time
        
        with lock:
            shared_dict['best_cost'] = best_cost
            shared_dict['elapsed_time'] = elapsed_time
            shared_dict['result_status'] = str(result)
            shared_dict['solver_stats'] = solver_stats
            shared_dict['timed_out'] = False
            shared_dict['completed'] = True
            shared_dict['model_count'] = model_count
        
    except Exception as e:
        print(f"Errore nel worker: {e}")
        with lock:
            shared_dict['best_cost'] = best_cost
            shared_dict['elapsed_time'] = time.time() - start_time
            shared_dict['result_status'] = "ERROR"
            shared_dict['solver_stats'] = {}
            shared_dict['timed_out'] = False
            shared_dict['completed'] = True
            shared_dict['model_count'] = model_count

def run_clingo_with_timeout(input_data, encoding, heuristics, timeout_seconds, control_args):

    manager = multiprocessing.Manager()
    shared_dict = manager.dict()
    lock = manager.Lock()
    
    shared_dict['best_cost'] = None
    shared_dict['model_count'] = 0
    shared_dict['elapsed_time'] = 0.0
    shared_dict['result_status'] = "UNKNOWN"
    shared_dict['solver_stats'] = {}
    shared_dict['timed_out'] = False
    shared_dict['completed'] = False
    
    start_time = time.time()
    
    worker = multiprocessing.Process(
        target=run_clingo_worker,
        args=(input_data, encoding, heuristics, control_args, timeout_seconds, shared_dict, lock)
    )
    worker.start()
    
    worker.join(timeout=timeout_seconds)
    
    if worker.is_alive():
        print(f"   *** TIMEOUT RAGGIUNTO DOPO {timeout_seconds} SECONDI - KILLING PROCESS ***")
        worker.terminate()
        worker.join(timeout=1)
        
        if worker.is_alive():
            worker.kill()
            worker.join()
        
        with lock:
            elapsed_time = time.time() - start_time
            result_data = {
                'elapsed_time': elapsed_time,
                'best_cost': shared_dict.get('best_cost', None),
                'model_count': shared_dict.get('model_count', 0),
                'result_status': "TIMEOUT",
                'solver_stats': shared_dict.get('solver_stats', {}),
                'timed_out': True
            }
        
        return result_data
    
    with lock:
        result_data = {
            'elapsed_time': shared_dict.get('elapsed_time', time.time() - start_time),
            'best_cost': shared_dict.get('best_cost', None),
            'model_count': shared_dict.get('model_count', 0),
            'result_status': shared_dict.get('result_status', "NO_RESULT"),
            'solver_stats': shared_dict.get('solver_stats', {}),
            'timed_out': False
        }
    
    return result_data

def find_all_input_files():
    input_dir = Path("input")
    if not input_dir.exists():
        print("Cartella input/ non trovata!")
        return []
    
    input_files = list(input_dir.rglob("*.lp"))
    input_files.sort(key=lambda x: str(x))
    
    return input_files

def run_multiple_times(input_data, encoding, heuristics, runs, timeout_seconds, control_args):

    results = []
    for i in range(runs):
        result = run_clingo_with_timeout(input_data, encoding, heuristics, timeout_seconds, control_args)
        results.append(result)
    
    times = [r['elapsed_time'] for r in results]
    valid_costs = [r['best_cost'] for r in results if r['best_cost'] is not None]
    model_counts = [r['model_count'] for r in results]
    timeouts = sum(1 for r in results if r['timed_out'])
    
    best_cost = None
    if valid_costs:
        best_cost = valid_costs[0]
        for cost in valid_costs[1:]:
            if is_cost_better(cost, best_cost):
                best_cost = cost
    
    stats = {
        'results': results,
        'avg_time': statistics.mean(times),
        'avg_models': statistics.mean(model_counts),
        'timeouts': timeouts,
        'total_runs': runs,
        'best_cost': best_cost,
        'valid_solutions': len(valid_costs)
    }
    
    return stats

def get_cost_components(cost):
    if cost is None:
        return []
    if isinstance(cost, list):
        return cost
    return [cost]

def calculate_speedup(original_time, config_time):
    if config_time == 0 or original_time == 0:
        return 0
    speedup = ((original_time - config_time) / original_time) * 100
    return round(speedup, 2)

def load_configuration_files():
    configurations = {}
    
    for i in range(1, 4):
        config_file = f"encoding_configuration_{i}.lp"
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                configurations[i] = f.read()
            print(f"Caricata configurazione {i} da: {config_file}")
        except FileNotFoundError:
            print(f"ATTENZIONE: file {config_file} non trovato!")
            configurations[i] = None
    
    return configurations

def create_comparison_dataframe(all_results):

    headers = [
        'Filename',
        'Cost 1 (Baseline)', 
        'Cost 2 (Baseline)',
        'Time Baseline',
        'Cost 1 (Cfg 1)',
        'Cost 2 (Cfg 1)', 
        'Time (Cfg 1)',
        'Speedup (Cfg 1)',
        'Cost 1 (Cfg 2)',
        'Cost 2 (Cfg 2)',
        'Time (Cfg 2)',
        'Speedup (Cfg 2)',
        'Cost 1 (Cfg 3)',
        'Cost 2 (Cfg 3)',
        'Time (Cfg 3)',
        'Speedup (Cfg 3)'
    ]
    
    data = []
    
    for result in all_results:
        input_name = Path(result['input_file']).name
        row = [f"original_encoding + {input_name}"]
        

        original_stats = result['original']
        original_cost = get_cost_components(original_stats['best_cost'])
        original_time = round(original_stats['avg_time'], 3)
        
        row.append(original_cost[0] if len(original_cost) > 0 else '')
        row.append(original_cost[1] if len(original_cost) > 1 else '')
        row.append(original_time)
        

        for config_num in [1, 2, 3]:
            if config_num in result and result[config_num] is not None:
                config_stats = result[config_num]
                config_cost = get_cost_components(config_stats['best_cost'])
                config_time = round(config_stats['avg_time'], 3)
                speedup = calculate_speedup(original_time, config_time)
                
                row.append(config_cost[0] if len(config_cost) > 0 else '')
                row.append(config_cost[1] if len(config_cost) > 1 else '')
                row.append(config_time)
                row.append(f"{speedup}%")
            else:

                row.extend(['', '', '', ''])
        
        data.append(row)
    
    df = pd.DataFrame(data, columns=headers)
    return df

def get_green_intensity(value, max_value, min_intensity=0.3):
    if max_value == 0:
        return min_intensity
    

    intensity = min_intensity + (value / max_value) * (1.0 - min_intensity)
    return min(intensity, 1.0)

def rgb_to_hex(r, g, b):
    r_int = int(r * 255)
    g_int = int(g * 255) 
    b_int = int(b * 255)
    return f"{r_int:02X}{g_int:02X}{b_int:02X}"

def write_results_to_excel(all_results, output_file):
    
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    df = create_comparison_dataframe(all_results)
    
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Configuration_Comparison"
        

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        

        max_speedup = 0
        max_time_improvement = 0
        baseline_times = []
        

        speedup_cols = [7, 11, 15]
        time_cols = [6, 10, 14]
        baseline_time_col = 3
        

        for row_idx in range(2, len(all_results) + 2):

            baseline_cell = ws.cell(row=row_idx, column=baseline_time_col)
            if baseline_cell.value and isinstance(baseline_cell.value, (int, float)):
                baseline_times.append(baseline_cell.value)
            

            for col_idx in speedup_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.endswith('%'):
                    try:
                        speedup_val = float(cell.value.replace('%', ''))
                        if speedup_val > max_speedup:
                            max_speedup = speedup_val
                    except ValueError:
                        pass
            

            if baseline_times:
                baseline_time = baseline_times[-1] if baseline_times else 0
                for col_idx in time_cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, (int, float)) and baseline_time > 0:
                        improvement_pct = ((baseline_time - cell.value) / baseline_time) * 100
                        if improvement_pct > max_time_improvement:
                            max_time_improvement = improvement_pct
        

        for row_idx in range(2, len(all_results) + 2):
            baseline_cell = ws.cell(row=row_idx, column=baseline_time_col)
            baseline_time = baseline_cell.value if baseline_cell.value and isinstance(baseline_cell.value, (int, float)) else 0
            

            for col_idx in speedup_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.endswith('%'):
                    try:
                        speedup_val = float(cell.value.replace('%', ''))
                        if speedup_val > 1.0:
                            intensity = get_green_intensity(speedup_val, max_speedup)
                            green_val = 0.5 + intensity * 0.5
                            color_hex = rgb_to_hex(1.0 - intensity * 0.3, green_val, 1.0 - intensity * 0.3)
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    except ValueError:
                        pass
            

            for col_idx in time_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, (int, float)) and baseline_time > 0:
                    if cell.value < baseline_time:
                        improvement_pct = ((baseline_time - cell.value) / baseline_time) * 100
                        intensity = get_green_intensity(improvement_pct, max_time_improvement)
                        green_val = 0.5 + intensity * 0.5
                        color_hex = rgb_to_hex(1.0 - intensity * 0.3, green_val, 1.0 - intensity * 0.3)
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        
        print(f"Risultati salvati in → {output_file}")
        print(f"Scritte {len(all_results)} righe di confronto")
        print(f"Formattazione condizionale applicata:")
        print(f"  - Speedup massimo trovato: {max_speedup:.1f}%")
        print(f"  - Miglioramento tempo massimo: {max_time_improvement:.1f}%")
        
    except ImportError:
        print("ERRORE: openpyxl non disponibile! Installa con: pip install openpyxl")

        csv_file = output_file.with_suffix('.csv')
        df.to_csv(csv_file, index=False)
        print(f"Salvato come CSV in → {csv_file}")
    except Exception as e:
        print(f"Errore nel salvataggio Excel: {e}")

        try:
            csv_file = output_file.with_suffix('.csv')
            df.to_csv(csv_file, index=False)
            print(f"Salvato come CSV in → {csv_file}")
        except Exception as csv_error:
            print(f"Errore anche nel salvataggio CSV: {csv_error}")

def run_configuration_test(input_file, base_encoding, configurations, config):
    
    print(f"=== PROCESSING: {input_file} ===")
    
    try:
        with open(input_file, "r", encoding="utf-8") as f:
            input_data = f.read()
    except Exception as e:
        print(f"Errore nel leggere {input_file}: {e}")
        return None
    
    result = {
        'input_file': str(input_file)
    }
    

    print("  → Test encoding originale...")
    original_stats = run_multiple_times(
        input_data,
        base_encoding,
        None,
        config["runs"],
        config["timeout_seconds"],
        config["control_args"]
    )
    result['original'] = original_stats
    print(f"    Tempo: {original_stats['avg_time']:.3f}s, Costo: {original_stats['best_cost']}")
    

    for config_num in [1, 2, 3]:
        if configurations[config_num] is not None:
            print(f"  → Test configurazione {config_num}...")
            config_stats = run_multiple_times(
                input_data,
                base_encoding,
                configurations[config_num],
                config["runs"],
                config["timeout_seconds"],
                config["control_args"]
            )
            result[config_num] = config_stats
            
            speedup = calculate_speedup(original_stats['avg_time'], config_stats['avg_time'])
            print(f"    Tempo: {config_stats['avg_time']:.3f}s, Costo: {config_stats['best_cost']}, Speedup: {speedup}%")
        else:
            print(f"  → Configurazione {config_num} non disponibile")
            result[config_num] = None
    
    print()
    return result

def main():
    config = load_config()
    
    print("=== RUN CONFIGURATION COMPARISON TESTS ===")
    print(f"Configurazione da settings.json:")
    print(f"  - Timeout: {config['timeout_seconds']} secondi")
    print(f"  - Runs per input: {config['runs']}")
    print(f"  - Base encoding: {config['base_encoding']}")
    print(f"  - Control args: {config['control_args']}")
    print()
    

    input_files = find_all_input_files()
    
    if not input_files:
        print("Nessun file .lp trovato nella cartella input/")
        return
    
    print(f"Trovati {len(input_files)} file input:")
    for f in input_files:
        print(f"  - {f}")
    print()
    

    try:
        with open(config["base_encoding"], "r", encoding="utf-8") as f:
            base_encoding = f.read()
    except FileNotFoundError:
        print(f"Errore: file encoding {config['base_encoding']} non trovato!")
        return
    

    configurations = load_configuration_files()
    print()
    

    all_results = []
    
    for i, input_file in enumerate(input_files, 1):
        print(f"=== PROCESSING {i}/{len(input_files)} ===")
        
        result = run_configuration_test(input_file, base_encoding, configurations, config)
        if result:
            all_results.append(result)
    

    if all_results:
        output_file = Path(config["timings_output_dir"]) / "configuration_comparison.xlsx"
        write_results_to_excel(all_results, output_file)
        

        print("=== RIASSUNTO FINALE ===")
        print(f"File processati: {len(all_results)}")
        
        for config_num in [1, 2, 3]:
            improvements = 0
            total_comparisons = 0
            
            for result in all_results:
                if config_num in result and result[config_num] is not None:
                    original_time = result['original']['avg_time']
                    config_time = result[config_num]['avg_time']
                    
                    if config_time < original_time:
                        improvements += 1
                    total_comparisons += 1
            
            if total_comparisons > 0:
                improvement_rate = (improvements / total_comparisons) * 100
                print(f"Configurazione {config_num}: {improvements}/{total_comparisons} miglioramenti ({improvement_rate:.1f}%)")
            else:
                print(f"Configurazione {config_num}: Non disponibile")
    
    print("\n=== COMPLETATO ===")

if __name__ == "__main__":
    multiprocessing.set_start_method('spawn', force=True)
    main()