
import clingo
import itertools
import statistics
import time
import re
import json
import multiprocessing
from pathlib import Path
import pandas as pd   

def load_config():
    settings_file = Path("settings.json")
    if not settings_file.exists():
        raise FileNotFoundError("File settings.json non trovato!")
    
    with settings_file.open(encoding="utf-8") as f:
        config = json.load(f)
    
    return config

CONFIG = load_config()

RUNS = CONFIG["runs"]
BASE_ENCODING = Path(CONFIG["base_encoding"])
HEURISTICS_FILE = Path(CONFIG["heuristics_file"])
INPUT_FILE = Path(CONFIG["input_file"])
CONTROL_ARGS = CONFIG["control_args"]
TIMEOUT_SECONDS = CONFIG.get("timeout_seconds", 30)
TIMINGS_OUTPUT_DIR = Path(CONFIG["timings_output_dir"])

def read(path: Path) -> str:
    with path.open(encoding="utf-8") as f:
        return f.read()

def get_timings_file(input_file: Path) -> Path:
    input_name = input_file.stem
    base_name = input_name
    
    timings_name = f"timings_{base_name}.xlsx"
    return TIMINGS_OUTPUT_DIR / timings_name

def cost_to_comparable(cost):
    if cost is None:
        return [float('inf')]
    if isinstance(cost, list):
        return cost
    return [cost]

def is_cost_better(cost1, cost2):
    c1 = cost_to_comparable(cost1)
    c2 = cost_to_comparable(cost2)
    
    for i in range(max(len(c1), len(c2))):
        val1 = c1[i] if i < len(c1) else 0
        val2 = c2[i] if i < len(c2) else 0
        
        if val1 < val2:
            return True
        elif val1 > val2:
            return False
    
    return False

def calculate_cost_improvement(baseline_cost, current_cost):
    baseline_components = cost_to_comparable(baseline_cost)
    current_components = cost_to_comparable(current_cost)
    
    if not baseline_components or baseline_components == [float('inf')]:
        return None
    
    def weighted_sum(components):
        total = 0
        for i, val in enumerate(components):
            if val == float('inf'):
                return float('inf')
            weight = 1000000 ** (len(components) - i - 1)
            total += val * weight
        return total
    
    baseline_weighted = weighted_sum(baseline_components)
    current_weighted = weighted_sum(current_components)
    
    if baseline_weighted == 0 or baseline_weighted == float('inf'):
        return None
    
    improvement = (baseline_weighted - current_weighted) / baseline_weighted
    return improvement

def run_clingo_worker(input_data, encoding, control_args, timeout_seconds, shared_dict, lock):
    best_cost = None
    model_count = 0
    start_time = time.time()
    
    def on_model(model):
        nonlocal best_cost, model_count
        model_count += 1
        current_cost = model.cost
        
        if best_cost is None or is_cost_better(current_cost, best_cost):
            best_cost = current_cost
            
            with lock:
                shared_dict['best_cost'] = current_cost
                shared_dict['model_count'] = model_count
                shared_dict['elapsed_time'] = time.time() - start_time
        
        elapsed = time.time() - start_time
        print(f"   Model {model_count}: Cost = {current_cost} (tempo: {elapsed:.2f}s)")
        return True
    
    try:
        ctl = clingo.Control(control_args)
        ctl.add("base", [], input_data)
        ctl.add("base", [], encoding)
        ctl.ground([("base", [])])
        
        result = ctl.solve(on_model=on_model)
        
        try:
            stats = ctl.statistics['solving']['solvers']
            solver_stats = {
                'choices': stats.get('choices', 0),
                'conflicts': stats.get('conflicts', 0)
            }
        except:
            solver_stats = {'choices': 0, 'conflicts': 0}
        
        elapsed_time = time.time() - start_time
        
        with lock:
            shared_dict['elapsed_time'] = elapsed_time
            shared_dict['result_status'] = str(result)
            shared_dict['solver_stats'] = solver_stats
            shared_dict['timed_out'] = False
            shared_dict['completed'] = True
        
    except Exception as e:
        print(f"   Errore nel worker: {e}")
        with lock:
            shared_dict['elapsed_time'] = time.time() - start_time
            shared_dict['result_status'] = "ERROR"
            shared_dict['solver_stats'] = {}
            shared_dict['timed_out'] = False
            shared_dict['completed'] = True

def run_clingo_with_timeout(input_data, encoding, timeout_seconds, control_args):
    
    manager = multiprocessing.Manager()
    shared_dict = manager.dict()
    lock = manager.Lock()
    
    shared_dict['best_cost'] = None
    shared_dict['model_count'] = 0
    shared_dict['elapsed_time'] = 0.0
    shared_dict['result_status'] = "UNKNOWN"
    shared_dict['solver_stats'] = {}
    shared_dict['timed_out'] = False
    shared_dict['completed'] = False
    
    start_time = time.time()
    
    worker = multiprocessing.Process(
        target=run_clingo_worker,
        args=(input_data, encoding, control_args, timeout_seconds, shared_dict, lock)
    )
    worker.start()
    
    worker.join(timeout=timeout_seconds)
    
    if worker.is_alive():
        print(f"   *** TIMEOUT RAGGIUNTO DOPO {timeout_seconds} SECONDI - KILLING PROCESS ***")
        worker.terminate()
        worker.join(timeout=1)
        
        if worker.is_alive():
            worker.kill()
            worker.join()
        
        with lock:
            elapsed_time = time.time() - start_time
            result_data = {
                'elapsed_time': elapsed_time,
                'best_cost': shared_dict.get('best_cost', None),
                'model_count': shared_dict.get('model_count', 0),
                'result_status': "TIMEOUT",
                'solver_stats': shared_dict.get('solver_stats', {}),
                'timed_out': True
            }
        
        return result_data
    
    with lock:
        result_data = {
            'elapsed_time': shared_dict.get('elapsed_time', time.time() - start_time),
            'best_cost': shared_dict.get('best_cost', None),
            'model_count': shared_dict.get('model_count', 0),
            'result_status': shared_dict.get('result_status', "NO_RESULT"),
            'solver_stats': shared_dict.get('solver_stats', {}),
            'timed_out': False
        }
    
    return result_data

def run_multiple_times(input_data: str, *encs: str) -> dict:
    
    results = []
    for i in range(RUNS):
        print(f"   → Run {i+1}/{RUNS}:", flush=True)
        
        combined_encoding = '\n'.join(encs)
        result = run_clingo_with_timeout(input_data, combined_encoding, TIMEOUT_SECONDS, CONTROL_ARGS)
        results.append(result)
        
        print(f"     Tempo: {result['elapsed_time']:.4f}s, Costo: {result['best_cost']}, "
              f"Modelli: {result['model_count']}, Status: {result['result_status']}")
        if result['timed_out']:
            print("     *** TIMEOUT CONFERMATO ***")
    
    times = [r['elapsed_time'] for r in results]
    valid_costs = [r['best_cost'] for r in results if r['best_cost'] is not None]
    model_counts = [r['model_count'] for r in results]
    timeouts = sum(1 for r in results if r['timed_out'])
    
    best_cost = None
    if valid_costs:
        best_cost = valid_costs[0]
        for cost in valid_costs[1:]:
            if is_cost_better(cost, best_cost):
                best_cost = cost
    
    stats = {
        'results': results,
        'avg_time': statistics.mean(times),
        'avg_models': statistics.mean(model_counts),
        'timeouts': timeouts,
        'total_runs': RUNS,
        'best_cost': best_cost,
        'valid_solutions': len(valid_costs)
    }
    
    print(f"   → Tempo medio: {stats['avg_time']:.4f}s")
    print(f"   → Modelli medi: {stats['avg_models']:.1f}")
    print(f"   → Timeout: {timeouts}/{RUNS}")
    if stats['best_cost'] is not None:
        print(f"   → Miglior costo: {stats['best_cost']}")
    print()
    
    return stats

def split_heuristics(text: str) -> dict[int, str]:
    blocks = {}
    lines = text.strip().splitlines()
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        match = re.match(r'^\s*%\s*Euristica\s*#(\d+)\s*$', line)
        if match:
            euristica_num = int(match.group(1))
            i += 1
            
            block_lines = []
            while i < len(lines):
                current_line = lines[i]
                if re.match(r'^\s*%\s*Euristica\s*#\d+\s*$', current_line.strip()):
                    break
                block_lines.append(current_line)
                i += 1
            
            heuristic_lines = []
            for bl in block_lines:
                stripped = bl.strip()
                if stripped and not stripped.startswith('%'):
                    heuristic_lines.append(bl)
            
            if heuristic_lines:
                blocks[euristica_num] = '\n'.join(heuristic_lines)
        else:
            i += 1
    
    return blocks

def get_cost_components(cost):
    if cost is None:
        return []
    if isinstance(cost, list):
        return cost
    return [cost]

def calculate_max_cost_components(all_results):
    max_components = 0
    for _, stats in all_results:
        if stats['best_cost'] is not None:
            components = get_cost_components(stats['best_cost'])
            max_components = max(max_components, len(components))
    return max_components

def create_fresh_dataframe(num_results: int, max_cost_components: int) -> pd.DataFrame:
    
    total_cols = 5 + max_cost_components
    
    total_rows = 8 + num_results
    
    column_names = list(range(total_cols))
    
    df = pd.DataFrame('', index=range(total_rows), columns=column_names)
    
    df.iloc[2, 1] = "Riassunto tempi e costi"
    df.iloc[4, 3] = f"Costi migliori\n(media su {RUNS} tentativi)"
    df.iloc[4, 3 + max_cost_components] = f"Tempo di esecuzione\n(media su {RUNS} tentativi)"
    df.iloc[4, 4 + max_cost_components] = "Speedup"
    
    df.iloc[6, 1] = "encoding"
    df.iloc[6, 2] = "euristica usata"
    
    for i in range(max_cost_components):
        col_letter = chr(ord('D') + i)
        df.iloc[6, 3 + i] = f"costo{i+1}"
    
    df.iloc[6, 3 + max_cost_components] = "tempo (s)"
    df.iloc[6, 4 + max_cost_components] = "speedup"
    
    return df

def sort_results_by_cost(results: list[tuple[list[int], dict]]) -> list[tuple[list[int], dict]]:
    def cost_sort_key(result_tuple):
        _, stats = result_tuple
        cost_components = get_cost_components(stats['best_cost'])
        
        if not cost_components:
            return [float('inf')]
        
        extended_components = cost_components + [0] * (10 - len(cost_components))
        return extended_components[:10]
    
    return sorted(results, key=cost_sort_key)

def write_to_timings(results: list[tuple[list[int], dict]], timings_file: Path) -> None:
    timings_file.parent.mkdir(parents=True, exist_ok=True)
    
    sorted_results = sort_results_by_cost(results)
    
    max_cost_components = calculate_max_cost_components(sorted_results)
    if max_cost_components == 0:
        max_cost_components = 1
    
    df = create_fresh_dataframe(len(sorted_results), max_cost_components)
    
    encoding_input = f"{BASE_ENCODING.stem} + {INPUT_FILE.stem}"
    
    current_row = 7
    
    baseline_time = None
    for ids, stats in sorted_results:
        if not ids:
            baseline_time = stats['avg_time']
            break
    
    for i, (ids, stats) in enumerate(sorted_results):
        if not ids:
            heuristics_str = "X"
        else:
            heuristics_str = " + ".join(map(str, ids))
        
        df.iloc[current_row, 0] = ''
        df.iloc[current_row, 1] = encoding_input
        df.iloc[current_row, 2] = heuristics_str
        
        cost_components = get_cost_components(stats['best_cost'])
        for j in range(max_cost_components):
            if j < len(cost_components):
                df.iloc[current_row, 3 + j] = cost_components[j]
            else:
                df.iloc[current_row, 3 + j] = ''
        
        df.iloc[current_row, 3 + max_cost_components] = round(stats['avg_time'], 3)
        
        if baseline_time is None or baseline_time == 0:
            df.iloc[current_row, 4 + max_cost_components] = ''
        else:
            excel_row = current_row + 1
            time_col = chr(ord('A') + 3 + max_cost_components)
            speedup_formula = f"=({baseline_time}-{time_col}{excel_row})/{baseline_time}"
            df.iloc[current_row, 4 + max_cost_components] = speedup_formula
        
        current_row += 1
    
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, numbers, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        
        with pd.ExcelWriter(timings_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
            
            ws = writer.sheets['Sheet1']
            
            d5_cell = ws['D5']
            d5_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            time_col = chr(ord('A') + 3 + max_cost_components)
            time_cell = ws[f'{time_col}5']
            time_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            speedup_col = chr(ord('A') + 4 + max_cost_components)
            
            for row in range(8, current_row + 1):
                cell = ws[f'{speedup_col}{row}']
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            
            if current_row > 7:
                speedup_range = f"{speedup_col}8:{speedup_col}{current_row}"
                rule = ColorScaleRule(
                    start_type='num', start_value=-0.5, start_color='FF0000',
                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                    end_type='num', end_value=0.5, end_color='00FF00'
                )
                ws.conditional_formatting.add(speedup_range, rule)
        
        print(f"Dati salvati in → {timings_file}")
        print(f"Scritte {len(sorted_results)} righe a partire da B8 (ordinate per costo)")
        print(f"Costi con {max_cost_components} componenti")
        
    except ImportError:
        print("ERRORE: openpyxl non disponibile! Installa con: pip install openpyxl")
    except Exception as e:
        print(f"Errore nel salvataggio: {e}")

def main() -> None:
    print("=== COMBINE HEURISTICS WITH TIMEOUT ===")
    print(f"Configurazione caricata da settings.json:")
    print(f"  - RUNS: {RUNS}")
    print(f"  - TIMEOUT: {TIMEOUT_SECONDS} secondi")
    print(f"  - BASE_ENCODING: {BASE_ENCODING}")
    print(f"  - INPUT_FILE: {INPUT_FILE}")
    print(f"  - HEURISTICS_FILE: {HEURISTICS_FILE}")
    
    timings_file = get_timings_file(INPUT_FILE)
    print(f"File timings: {timings_file}")
    print()
    
    inp_text      = read(INPUT_FILE)
    base_enc_text = read(BASE_ENCODING)
    
    if not HEURISTICS_FILE.exists():
        print(f"File {HEURISTICS_FILE} non trovato!")
        return
    
    heur_text = read(HEURISTICS_FILE)
    heur_blocks = split_heuristics(heur_text)

    results = []

    print("Calcolo baseline…")
    baseline_stats = run_multiple_times(inp_text, base_enc_text)
    print(f"Baseline: {baseline_stats['avg_time']:.3f}s (media su {RUNS} run)")
    if baseline_stats['best_cost'] is not None:
        print(f"Miglior costo baseline: {baseline_stats['best_cost']}")
    results.append(([], baseline_stats))

    if heur_blocks:
        print(f"Trovate {len(heur_blocks)} euristiche promettenti")
        
        for euristica_id in sorted(heur_blocks.keys()):
            print(f"Test euristica {euristica_id}…")
            heuristic_stats = run_multiple_times(inp_text, base_enc_text, heur_blocks[euristica_id])
            print(f"Euristica {euristica_id}: {heuristic_stats['avg_time']:.3f}s")
            if heuristic_stats['best_cost'] is not None:
                print(f"Miglior costo: {heuristic_stats['best_cost']}")
            results.append(([euristica_id], heuristic_stats))

        if len(heur_blocks) > 1:
            for k in range(2, len(heur_blocks) + 1):
                print(f"Combinazioni di {k} euristiche…")
                for combo in itertools.combinations(sorted(heur_blocks.keys()), k):
                    combo_texts = [heur_blocks[i] for i in combo]
                    print(f"Test combinazione {' + '.join(map(str, combo))}…")
                    combo_stats = run_multiple_times(inp_text, base_enc_text, *combo_texts)
                    print(f"  {' + '.join(map(str, combo))}: {combo_stats['avg_time']:.3f}s")
                    if combo_stats['best_cost'] is not None:
                        print(f"  Miglior costo: {combo_stats['best_cost']}")
                    results.append((list(combo), combo_stats))
    else:
        print("Nessuna euristica trovata in 2_promising_ones.lp")

    write_to_timings(results, timings_file)

if __name__ == "__main__":
    multiprocessing.set_start_method('spawn', force=True)
    main()
